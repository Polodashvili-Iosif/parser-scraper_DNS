import csv
import os
import xml.etree.ElementTree as ElementTree
from datetime import datetime
from json import dumps
from typing import Any, Iterable, Mapping, Optional
from xml.dom.minidom import parseString

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

if not os.path.exists('resulting files'):
    os.makedirs('resulting files')


def get_date_and_time() -> str:
    return datetime.now().strftime('%d.%m.%y %H-%M-%S')


def to_xml(data: Iterable[Mapping[str, Any]],
           parameters: Optional[list[str]] = None,
           indent: Optional[str] = '\t', root: Optional[str] = 'items',
           item_name: Optional[str] = 'item',
           file_name: Optional[str] = 'data') -> None:
    """
    Создаёт из итерируемого объекта xml файл в папке "resulting files."
    """
    root = ElementTree.Element(root)
    for item in data:
        item_root = ElementTree.Element(item_name)
        root.append(item_root)
        if parameters:
            for name in parameters:
                item_param = ElementTree.SubElement(
                    item_root, name.replace(' ', '_'))
                item_param.text = str(item.get(name, ''))
        else:
            for name, value in item.items():
                item_param = ElementTree.SubElement(
                    item_root, name.replace(' ', '_'))
                item_param.text = str(value)

    tree = ElementTree.ElementTree(root)

    text = parseString(ElementTree.tostring(
        tree.getroot(), encoding='utf-8', method='xml'
    )).toprettyxml(encoding='utf-8', indent=indent)

    datetime_now = get_date_and_time()
    with open(f"resulting files/{file_name} {datetime_now}.xml", "wb") as file:
        file.write(text)


def to_json(data: Iterable[Mapping[str, Any]],
            file_name: Optional[str] = "data") -> None:
    """
    Создаёт из итерируемого объекта json файл в папке "resulting files".
    """
    datetime_now = get_date_and_time()
    with open(f'resulting files/{file_name} {datetime_now}.json', 'w', encoding='utf-8') as file:
        file.write(dumps(data, indent=4, ensure_ascii=False))


def to_excel(data: Iterable[Mapping[str, Any]], column_names: Iterable[str],
             file_name: Optional[str] = "table") -> None:
    """
    Создаёт из итерируемого объекта и имён столбцов
    xlsx файл в папке "resulting files".
    """
    wb = Workbook()
    worksheet = wb.active
    side = Side(border_style='thin')
    border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side
    )
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    )
    column_widths = []

    for column, name in enumerate(column_names, 1):
        cell = worksheet.cell(
            column=column,
            row=1,
            value=name
        )
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = alignment
        column_widths.append(len(name))

    for row, product in enumerate(data, 2):
        if not product:
            print(row)
            continue
        for column, name in enumerate(column_names, 1):
            cell = worksheet.cell(
                column=column,
                row=row,
                value=product.get(name, 'Нет')
            )
            cell.border = border
            cell.alignment = alignment
            column_widths[column -
                          1] = max(column_widths[column -
                                                 1], len(str(cell.value)))

    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(
            i)].width = column_width + 1

    datetime_now = get_date_and_time()
    wb.save(f"resulting files/{file_name} {datetime_now}.xlsx")


def to_csv(data: Iterable[Mapping[str, Any]], column_names: list[str],
           file_name: Optional[str] = 'data') -> None:
    """
    Создаёт из итерируемого объекта и имён столбцов
    csv файл в папке "resulting files".
    """
    datetime_now = get_date_and_time()
    with open(f'resulting files/{file_name} {datetime_now}.csv', 'w', encoding='utf-8', newline='') as file:
        writer = csv.DictWriter(file, column_names)
        writer.writeheader()
        for item in data:
            writer.writerow(item)
